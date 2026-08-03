"""
Spreadsheet (.xlsx) document loader.

Converts each worksheet of an Office Open XML workbook into a plain-text
document where every row becomes a pipe-separated line, so the regular
chunking/embedding pipeline can process it like any other text document.
"""

import os
from typing import List

from langchain_core.documents import Document

from ..logging_config import get_logger
from .document_processing_types import DocumentList

logger = get_logger(__name__)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


class SpreadsheetLoader:
    """Loader that turns an .xlsx workbook into one Document per worksheet."""

    def __init__(self, file_path: str) -> None:
        self.file_path = file_path

    def load(self) -> DocumentList:
        from openpyxl import load_workbook

        file_name = os.path.basename(self.file_path)
        workbook = load_workbook(filename=self.file_path, read_only=True, data_only=True)
        documents: List[Document] = []

        try:
            for sheet in workbook.worksheets:
                lines: List[str] = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [_format_cell(value) for value in row]
                    if not any(cells):
                        continue
                    lines.append(" | ".join(cells))

                if not lines:
                    continue

                content = f"# {sheet.title}\n" + "\n".join(lines)
                documents.append(
                    Document(
                        page_content=content,
                        metadata={
                            "source": self.file_path,
                            "sheet_name": sheet.title,
                            "row_count": len(lines),
                        },
                    )
                )
        finally:
            workbook.close()

        logger.debug(
            f"SpreadsheetLoader extracted {len(documents)} sheet(s) from {file_name}"
        )
        return documents
